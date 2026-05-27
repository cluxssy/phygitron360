import io
import calendar
from typing import List, Dict, Any, Optional
from backend.modules.deploy.repositories.payroll_repo import PayrollRepository
from backend.modules.deploy.services.notification_service import add_notification

MONTH_NAMES = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December'
}


class PayrollService:
    def __init__(self, tenant_id: str = 'public'):
        self.repo = PayrollRepository()
        self.tenant_id = tenant_id

    def parse_excel(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Parse a multi-sheet xlsx. Each sheet name = employee_code."""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        records = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            employee_code = sheet_name.strip()
            data = {'employee_code': employee_code}
            label_map = {
                'basic salary': 'basic_salary',
                'hra': 'hra',
                'special allowance': 'special_allowance',
                'medical insurance': 'medical_insurance',
                "pf employer's contribution": 'pf_employer_contribution',
                'pf employer contribution': 'pf_employer_contribution',
                'travelling reimbursement': 'travelling_reimbursement',
                'income tax': 'income_tax',
                'medical': 'medical_deduction',
                'employer pf': 'employer_pf',
                'employee pf': 'employee_pf',
            }
            deduction_section = False
            for row in ws.iter_rows(values_only=True):
                # Find non-None cells
                vals = [v for v in row if v is not None]
                if len(vals) < 2:
                    continue
                label = str(vals[0]).strip().lower()
                # Detect section switch
                if label == 'deduction':
                    deduction_section = True
                    continue
                if label == 'ctc':
                    ctc_vals = [v for v in row if isinstance(v, (int, float)) and v > 0]
                    if len(ctc_vals) > 1:
                        # Grab the Annual value (second column)
                        data['gross_ctc'] = float(ctc_vals[1])
                    elif len(ctc_vals) > 0:
                        data['gross_ctc'] = float(ctc_vals[0])
                    continue
                if label == 'total' and deduction_section:
                    for v in row:
                        if isinstance(v, (int, float)) and v > 0:
                            data['total_deductions'] = float(v)
                            break
                    continue
                if label == 'net in hand':
                    for v in row:
                        if isinstance(v, (int, float)) and v > 0:
                            data['net_in_hand'] = float(v)
                            break
                    continue
                # Map field names
                mapped = label_map.get(label)
                if mapped:
                    for v in row:
                        if isinstance(v, (int, float)) and v > 0:
                            data[mapped] = float(v)
                            break
            
            earnings_keys = [
                'basic_salary', 'hra', 'special_allowance', 'medical_insurance',
                'pf_employer_contribution', 'travelling_reimbursement'
            ]
            monthly_earnings = sum(data.get(k, 0) for k in earnings_keys)
            data['monthly_ctc'] = monthly_earnings
            
            ded_keys = ['income_tax', 'medical_deduction', 'employer_pf', 'employee_pf']

            # Compute fallbacks if formula wasn't resolved
            if 'gross_ctc' not in data:
                data['gross_ctc'] = monthly_earnings * 12
            if 'total_deductions' not in data:
                data['total_deductions'] = sum(data.get(k, 0) for k in ded_keys)
            if 'net_in_hand' not in data:
                data['net_in_hand'] = monthly_earnings - data.get('total_deductions', 0)
                
            # Fetch employee info to append for preview
            emp = self.repo.get_employee_info(employee_code, self.tenant_id)
            if emp:
                data['employee_name'] = emp.get('name')
                data['bank_name'] = emp.get('bank_name')
                data['bank_account_no'] = emp.get('bank_account_no')
                data['pan_no'] = emp.get('pan_no')
                data['designation'] = emp.get('designation')
                data['team'] = emp.get('team')
                data['location'] = emp.get('location')
                
            records.append(data)
        return records

    def push_pay_cycle(self, records: List[Dict], pay_month: int, pay_year: int,
                       pay_date: Optional[str], uploaded_by: str) -> Dict:
        """Save all records to DB and notify employees."""
        success_count = 0
        error_codes = []
        for rec in records:
            try:
                rec['pay_month'] = pay_month
                rec['pay_year'] = pay_year
                rec['pay_date'] = pay_date
                rec['uploaded_by'] = uploaded_by
                self.repo.upsert_payroll_record(rec, self.tenant_id)
                # Send notification to employee
                add_notification(
                    title=f"Payslip Available \u2014 {MONTH_NAMES[pay_month]} {pay_year}",
                    message=(
                        f"Your payslip for {MONTH_NAMES[pay_month]} {pay_year} has been processed. "
                        f"Net Pay: \u20b9{rec.get('net_in_hand', 0):,.2f}"
                    ),
                    employee_code=rec['employee_code'],
                    n_type="Success",
                    tenant_id=self.tenant_id
                )
                success_count += 1
            except Exception as e:
                error_codes.append(rec.get('employee_code', 'unknown'))
        return {'success_count': success_count, 'errors': error_codes}

    def get_employee_payslips(self, employee_code: str) -> List[Dict]:
        return self.repo.get_employee_payslips(employee_code, self.tenant_id)

    def get_payroll_record(self, employee_code: str, pay_month: int, pay_year: int) -> Optional[Dict]:
        return self.repo.get_payroll_record(employee_code, pay_month, pay_year, self.tenant_id)

    def get_cycle_summary(self, pay_month: int, pay_year: int) -> List[Dict]:
        return self.repo.get_cycle_payslips(pay_month, pay_year, self.tenant_id)

    def get_distinct_cycles(self) -> List[Dict]:
        return self.repo.get_distinct_cycles(self.tenant_id)

    def generate_payslip_pdf(self, employee_code: str, pay_month: int, pay_year: int) -> Optional[bytes]:
        """Generate a PDF payslip strictly matching the provided docx template."""
        record = self.repo.get_payroll_record(employee_code, pay_month, pay_year, self.tenant_id)
        if not record:
            return None
        emp = self.repo.get_employee_info(employee_code, self.tenant_id) or {}
        # Merge emp info into record for the generic generator
        for k, v in emp.items():
            if k not in record:
                record[k] = v
        return self._generate_pdf_from_record(employee_code, pay_month, pay_year, record)

    def generate_preview_payslip_pdf(self, record: dict) -> Optional[bytes]:
        """Generate a PDF preview from a raw JSON payload (without fetching from DB)."""
        employee_code = record.get('employee_code', 'Unknown')
        pay_month = record.get('pay_month', 1)
        pay_year = record.get('pay_year', 2025)
        return self._generate_pdf_from_record(employee_code, pay_month, pay_year, record)

    def _generate_pdf_from_record(self, employee_code: str, pay_month: int, pay_year: int, record: dict) -> bytes:
        import os
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, HRFlowable
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        emp = record
        buf = io.BytesIO()
        
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.8 * inch,
            bottomMargin=0.8 * inch
        )
        
        elements = []

        # --- Header ---
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
        logo_path = os.path.join(base_dir, 'assets', 'Picture 1.png')
        
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=1.2*inch, height=1.2*inch)
            logo.hAlign = 'LEFT'
        else:
            logo = ''

        company_info = [
            Paragraph('<b>ewandzdigital Services Pvt. Ltd.</b>', ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=18, alignment=TA_CENTER, leading=22)),
            Paragraph('20, Okhla Phase 3, New Delhi, Delhi - 110020', ParagraphStyle('Address', fontName='Helvetica', fontSize=10, alignment=TA_CENTER, leading=14)),
            Paragraph('CIN: U72900DL2017PTC327055', ParagraphStyle('CIN', fontName='Helvetica', fontSize=10, alignment=TA_CENTER, leading=14))
        ]
        
        # Use a 3-column table to force the company info into the absolute center of the page
        # while keeping the logo on the far left. (Total width = 7.27 inches)
        header_table = Table([[logo, company_info, '']], colWidths=[1.5*inch, 4.27*inch, 1.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 20))

        month_name = MONTH_NAMES.get(pay_month, str(pay_month))
        elements.append(Paragraph(f'Payslip for the Month of {month_name}, {pay_year}', ParagraphStyle('Month', fontName='Helvetica', fontSize=14, alignment=TA_CENTER)))
        elements.append(Spacer(1, 10))

        # --- Table Data ---
        col_w = [1.5 * inch, 2.135 * inch, 1.5 * inch, 2.135 * inch]
        
        emp_name = emp.get('name') or emp.get('employee_name') or ''
        t_data = [
            ['Name:', emp_name, 'Employee ID:', employee_code],
            ['Designation:', emp.get('designation', ''), 'Bank Name:', emp.get('bank_name', '')],
            ['Department:', emp.get('team', ''), 'Bank Account No.:', emp.get('bank_account_no', '')],
            ['Location:', emp.get('location', ''), 'PAN No.:', emp.get('pan_no', '')],
            ['Effective Work Days:', '', '', ''],
            ['LOP:', '', '', ''],
            ['Earnings', 'Amount', 'Deductions', 'Amount']
        ]

        earn_items = [
            ('Basic', record.get('basic_salary', 0)),
            ('HRA', record.get('hra', 0)),
            ('Leave Travel Allowance', 0), # Added based on template
            ('Special Allowance', record.get('special_allowance', 0)),
            ('Medical Insurance', record.get('medical_insurance', 0)),
            ('PF Employer Contribution', record.get('pf_employer_contribution', 0)),
            ('Travelling Reimbursement', record.get('travelling_reimbursement', 0)),
        ]
        # Filter out 0s for some except core ones if needed, but we'll show all
        
        ded_items = [
            ('PF Employee', record.get('employee_pf', 0)),
            ('Employer PF', record.get('employer_pf', 0)),
            ('Income Tax', record.get('income_tax', 0)),
            ('Medical', record.get('medical_deduction', 0)),
        ]

        max_len = max(len(earn_items), len(ded_items))
        
        def fmt(v):
            if isinstance(v, (int, float)) and v != 0:
                return f'{float(v):,.2f}'
            elif v == 0:
                return '0'
            return str(v)

        total_earn = sum(float(val) for _, val in earn_items)
        total_ded = sum(float(val) for _, val in ded_items)

        for i in range(max_len):
            e_label = earn_items[i][0] if i < len(earn_items) else ''
            e_val   = fmt(earn_items[i][1]) if i < len(earn_items) else ''
            d_label = ded_items[i][0] if i < len(ded_items) else ''
            d_val   = fmt(ded_items[i][1]) if i < len(ded_items) else ''
            t_data.append([e_label, e_val, d_label, d_val])

        t_data.append([
            'Total Earnings (Rs)', fmt(total_earn),
            'Total Deductions (Rs)', fmt(total_ded)
        ])

        t = Table(t_data, colWidths=col_w)
        
        ts = TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
            ('LINEBEFORE', (2, 0), (2, -1), 1.5, colors.black),
            
            ('LINEABOVE', (0, 6), (-1, 6), 1.5, colors.black),
            ('LINEBELOW', (0, 6), (-1, 6), 1.5, colors.black),
            
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.black),
            
            ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            
            ('ALIGN', (0, 0), (1, 5), 'LEFT'),
            ('ALIGN', (2, 0), (3, 5), 'LEFT'),
            
            ('ALIGN', (0, 6), (0, 6), 'LEFT'),
            ('ALIGN', (1, 6), (1, 6), 'RIGHT'),
            ('ALIGN', (2, 6), (2, 6), 'LEFT'),
            ('ALIGN', (3, 6), (3, 6), 'RIGHT'),
            
            ('ALIGN', (0, 7), (0, -1), 'LEFT'),
            ('ALIGN', (1, 7), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 7), (2, -1), 'LEFT'),
            ('ALIGN', (3, 7), (3, -1), 'RIGHT'),
            
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ])
        
        t.setStyle(ts)
        elements.append(t)
        
        elements.append(Spacer(1, 15))
        
        elements.append(Paragraph(
            f'<b>Net Pay For The Month:</b>', 
            ParagraphStyle('NetLabel', fontName='Helvetica', fontSize=9, alignment=TA_LEFT, spaceAfter=2)
        ))
        elements.append(Paragraph(
            '(Rupees Only)', 
            ParagraphStyle('sm', fontName='Helvetica', fontSize=9, alignment=TA_LEFT)
        ))
        
        elements.append(Spacer(1, 40))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.black, spaceAfter=15))
        
        elements.append(Paragraph(
            'This is a system generated payslip and does not require signature.',
            ParagraphStyle('Footer', fontName='Helvetica', fontSize=10, alignment=TA_CENTER)
        ))

        doc.build(elements)
        buf.seek(0)
        return buf.read()
